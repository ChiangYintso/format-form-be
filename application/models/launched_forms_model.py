from bson import ObjectId
from openpyxl import Workbook

from .person_model import PersonModel
from application.utils.exception.custom_exception import CustomException


class LaunchedFormsModel:

    def __init__(self, current_app, open_id):
        self.__current_app = current_app
        self.__person = PersonModel(self.__current_app, open_id)

    def put_launched_form_status(
            self, form_temp_id: str, form_type: str, end_time: str):
        """
        This method is used for updating the form status in mini-program
        detail page. If you need to update the whole form's data, use method
        `put_launched_forms`.

        :param form_temp_id: ObjectId of forms.
        :param form_type: Type of form, values 'ended' or 'launched'.
        :param end_time: The end time of the form.
        :raise: CustomException if no matched document in mongodb.
        """

        res = self.__current_app.mongo.db.form_templates.update_one({
            '_id': ObjectId(form_temp_id)
        }, {
            '$set': {'type': form_type, 'end_time': end_time}
        })

        if res.matched_count <= 0:
            raise CustomException(3000, 'invalid data', 300)

    def put_launched_form(self, data):
        res = self.__current_app.mongo.db.form_templates.update_one({
            '_id': ObjectId(data['form_temp_id'])
        }, {
            '$set': {
                'title': data['title'],
                'questions': data['questions'],
                'type': data['type'],
                'show_select_res': data['show_select_res'],
                'repeat_filling': data['repeat_filling'],
                'start_time': data['start_time'],
                'end_time': data['end_time']
            }
        })

        if res.matched_count <= 0:
            raise CustomException(3000, 'invalid data', 300)

    def generate_excel(self, form_id):
        res = self.__current_app.mongo.db.form_templates.find_one({
            '_id': ObjectId(form_id),
            'open_id': self.__person.get_open_id()
        }, projection={
            'title': True,
            'questions': True,
            'form_data': True
        })

        if res is None:
            raise CustomException(3000, 'invalid data', 200)

        wb = Workbook()

        ws1 = wb.active
        ws1.merge_cells('A1:D1')
        ws1.cell(1, 1).value = res['title']

        ws1.append([item['desc'] for item in res['questions']])

        for people in res['form_data']:
            row = []
            order = 0
            for ans in people['answer']:
                if res['questions'][order]['type'] == 'radio':
                    row.append(res['questions'][order]['detail'][int(ans)])
                elif res['questions'][order]['type'] == 'essay':
                    row.append(ans)
                else:
                    s = ''
                    for option in ans:
                        s += res['questions'][order]['detail'][int(option)]+';'
                    row.append(s)
                order += 1
            ws1.append(row)

        dirname = './application/static/excel/{}.xlsx'.format(form_id)
        wb.save(filename=dirname)
